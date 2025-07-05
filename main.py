import sqlite3
from fastapi import FastAPI, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from typing import List
from datetime import datetime
from fastapi.responses import FileResponse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse
import io

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # O especifica tu frontend: ["http://localhost:5500"]
    allow_methods=["*"],
    allow_headers=["*"],
)


DB_FILE = "asistencia.db"
EXCEL_FILE = "asistencia.xlsx"


def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def crear_base_datos():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS personas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS asistencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            persona_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            estado TEXT NOT NULL,
            FOREIGN KEY(persona_id) REFERENCES personas(id)
        )
    """)
    conn.commit()
    conn.close()

crear_base_datos()

@app.get("/nombres", response_model=List[str])
def obtener_nombres():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT nombre FROM personas ORDER BY nombre")
    nombres = [row["nombre"] for row in cursor.fetchall()]
    conn.close()
    return nombres

@app.post("/agregar")
def agregar_persona(nombre: str = Form(...)):
    nombre = nombre.strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO personas (nombre) VALUES (?)", (nombre,))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="La persona ya existe")
    conn.close()
    return {"mensaje": "Persona agregada correctamente"}

@app.post("/marcar")
def marcar_asistencia(nombre: str = Form(...), estado: str = Form(...), fecha: str = Form(None)):
    nombre = nombre.strip()
    if fecha:
        try:
            fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
            fecha_str = fecha_dt.strftime("%Y-%m-%d")  # Guardar en ISO para consistencia
        except:
            raise HTTPException(status_code=400, detail="Fecha inválida")
    else:
        fecha_str = datetime.now().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM personas WHERE nombre = ?", (nombre,))
    persona = cursor.fetchone()
    if not persona:
        conn.close()
        raise HTTPException(status_code=404, detail="La persona no está registrada")

    persona_id = persona["id"]

    # Si ya hay registro para esa persona y fecha, actualizamos, si no insertamos
    cursor.execute("""
        SELECT id FROM asistencias WHERE persona_id = ? AND fecha = ?
    """, (persona_id, fecha_str))
    registro = cursor.fetchone()

    if registro:
        cursor.execute("""
            UPDATE asistencias SET estado = ? WHERE id = ?
        """, (estado, registro["id"]))
    else:
        cursor.execute("""
            INSERT INTO asistencias (persona_id, fecha, estado) VALUES (?, ?, ?)
        """, (persona_id, fecha_str, estado))

    conn.commit()
    conn.close()
    return {"mensaje": f"Asistencia registrada para {fecha_str}"}

@app.put("/editar")
def editar_persona(nombre_actual: str = Form(...), nuevo_nombre: str = Form(...)):
    nombre_actual = nombre_actual.strip()
    nuevo_nombre = nuevo_nombre.strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM personas WHERE nombre = ?", (nombre_actual,))
    persona = cursor.fetchone()
    if not persona:
        conn.close()
        raise HTTPException(status_code=404, detail="Persona no encontrada")

    try:
        cursor.execute("UPDATE personas SET nombre = ? WHERE id = ?", (nuevo_nombre, persona["id"]))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="El nuevo nombre ya existe")
    conn.close()
    return {"mensaje": "Nombre actualizado"}

@app.delete("/borrar")
def borrar_persona(nombre: str = Form(...)):
    nombre = nombre.strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM personas WHERE nombre = ?", (nombre,))
    persona = cursor.fetchone()
    if not persona:
        conn.close()
        raise HTTPException(status_code=404, detail="Persona no encontrada")

    cursor.execute("DELETE FROM asistencias WHERE persona_id = ?", (persona["id"],))
    cursor.execute("DELETE FROM personas WHERE id = ?", (persona["id"],))
    conn.commit()
    conn.close()
    return {"mensaje": "Persona eliminada"}

@app.get("/resumen/personas")
def resumen_por_persona():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT p.nombre,
            SUM(CASE WHEN a.estado = 'Asistencia' THEN 1 ELSE 0 END) AS asistencias,
            SUM(CASE WHEN a.estado = 'Falta' THEN 1 ELSE 0 END) AS faltas
        FROM personas p
        LEFT JOIN asistencias a ON p.id = a.persona_id
        GROUP BY p.id, p.nombre
        ORDER BY p.nombre
    """)

    resultados = cursor.fetchall()
    conn.close()

    resumen = [{"nombre": row["nombre"], "asistencias": row["asistencias"] or 0, "faltas": row["faltas"] or 0} for row in resultados]
    return resumen

@app.get("/resumen/personas/mes")
def resumen_personas_mes(mes: str):  # mes formato "YYYY-MM"
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT p.nombre,
            SUM(CASE WHEN a.estado = 'Asistencia' AND strftime('%Y-%m', a.fecha) = ? THEN 1 ELSE 0 END) AS asistencias,
            SUM(CASE WHEN a.estado = 'Falta' AND strftime('%Y-%m', a.fecha) = ? THEN 1 ELSE 0 END) AS faltas
        FROM personas p
        LEFT JOIN asistencias a ON p.id = a.persona_id
        GROUP BY p.id, p.nombre
        ORDER BY p.nombre
    """, (mes, mes))

    resultados = cursor.fetchall()
    conn.close()

    resumen = [{"nombre": row["nombre"], "asistencias": row["asistencias"] or 0, "faltas": row["faltas"] or 0} for row in resultados]
    return resumen

@app.get("/resumen/dias")
def resumen_por_dia():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT fecha,
            SUM(CASE WHEN estado = 'Asistencia' THEN 1 ELSE 0 END) AS asistencias
        FROM asistencias
        GROUP BY fecha
        ORDER BY fecha
    """)

    resultados = cursor.fetchall()
    conn.close()

    resumen = [{"fecha": row["fecha"], "asistencias": row["asistencias"] or 0} for row in resultados]
    return resumen

@app.get("/descargar-excel")
def descargar_excel():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener nombres y fechas únicas
    cursor.execute("SELECT nombre FROM personas ORDER BY nombre")
    personas = [row["nombre"] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT fecha FROM asistencias ORDER BY fecha")
    fechas = [row["fecha"] for row in cursor.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    # Estilos base
    font_bold = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fill_even_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Estilos por estado
    fill_asistencia = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde claro
    fill_falta = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # rojo claro
    fill_vacio = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")      # amarillo claro

    # Agregar encabezado
    encabezado = ["Nombre"] + fechas + ["Total Asistencias", "Total Faltas"]
    ws.append(encabezado)

    for col in ws[1]:
        col.font = font_bold
        col.fill = fill_header
        col.alignment = align_center
        col.border = border

    # Llenar datos por persona
    for idx, persona in enumerate(personas, start=2):
        fila = [persona]
        asistencias = 0
        faltas = 0

        for fecha in fechas:
            cursor.execute("""
                SELECT estado FROM asistencias a
                JOIN personas p ON a.persona_id = p.id
                WHERE p.nombre = ? AND a.fecha = ?
            """, (persona, fecha))
            resultado = cursor.fetchone()
            estado = resultado["estado"] if resultado else ""
            fila.append(estado)
            if estado == "Asistencia":
                asistencias += 1
            elif estado == "Falta":
                faltas += 1

        fila += [asistencias, faltas]
        ws.append(fila)

        # Aplicar estilos a la fila
        for j, celda in enumerate(ws[idx], start=1):
            celda.alignment = align_center
            celda.border = border

            # Colorear asistencias/faltas según estado
            if 1 < j <= len(fechas) + 1:  # columnas de fechas
                if celda.value == "Asistencia":
                    celda.fill = fill_asistencia
                elif celda.value == "Falta":
                    celda.fill = fill_falta
                else:
                    celda.fill = fill_vacio
            elif idx % 2 == 0:
                celda.fill = fill_even_row  # fondo alterno para filas

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_length + 2, 12)

    # Guardar archivo en memoria
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    conn.close()

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=asistencias.xlsx"}
    )